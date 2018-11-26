# _*_ coding:utf-8 _*_
import openpyxl
import xlrd
import os


# 写入到一个excel
def write_excel(list):
    out_wb = openpyxl.Workbook()  # 打开一个将写的文件
    out_ws = out_wb.create_sheet(index=0)  # 在将写的文件创建sheet

    for row in range(len(list)):
        print("正在导入：%s" % list[row])
        for col in range(len(list[row])):
            out_ws.cell(row + 1, col + 1).value = list[row][col].value  # 写文件
    total_execl_path = r"C:\Users\pig\Desktop\total.xlsx"
    out_wb.save(total_execl_path)  # 一定要记得保存


if __name__ == "__main__":

    data_path = r'C:\Users\pig\Desktop\data_v1.1'
    all_data = []

    for dir_path, dir_names, file_names in os.walk(data_path):
        for file_path in file_names:
            # excel文件路径
            excel_path = os.path.join(dir_path, file_path)
            print(excel_path)
            # 打开excel
            excel_now = xlrd.open_workbook(excel_path)
            # 获取第一个sheet内容
            sheet_content = excel_now.sheet_by_index(0)
            # 获取最大行数和列数
            rows = sheet_content.nrows
            cols = sheet_content.ncols
            for i in range(1, rows):
                rows_temp = []
                for j in range(0, cols):
                    rows_temp.append(sheet_content.cell(i, j))
                print(rows_temp)
                all_data.append(rows_temp)
    write_excel(all_data)
